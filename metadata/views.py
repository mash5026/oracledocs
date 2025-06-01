from django.shortcuts import render
from django.db import connection
import openpyxl
from django.http import HttpResponse
from openpyxl.utils import get_column_letter


def index(request):
    return render(request, "metadata/index.html")

def get_schema_list():
    with connection.cursor() as cursor:
        cursor.execute("SELECT DISTINCT OWNER FROM ALL_OBJECTS ORDER BY OWNER")
        return [row[0] for row in cursor.fetchall()]

def table_metadata(request):

    owner = request.GET.get("owner")
    export = request.GET.get("export")

    query = """
        SELECT t.OWNER,
               t.TABLE_NAME,
               t.COLUMN_NAME,
               t.DATA_TYPE,
               t.DATA_LENGTH,
               t.NULLABLE,
               t.DATA_DEFAULT,
               CASE WHEN pk_cols.COLUMN_NAME IS NOT NULL THEN 'YES' ELSE 'NO' END AS IS_PRIMARY_KEY,
               idx.INDEX_NAME,
               tab.NUM_ROWS,
               c.COMMENTS
        FROM ALL_TAB_COLUMNS t
        LEFT JOIN ALL_COL_COMMENTS c
            ON t.OWNER = c.OWNER AND t.TABLE_NAME = c.TABLE_NAME AND t.COLUMN_NAME = c.COLUMN_NAME
        LEFT JOIN (
            SELECT a.OWNER, a.TABLE_NAME, a.COLUMN_NAME
            FROM ALL_CONS_COLUMNS a
            JOIN ALL_CONSTRAINTS c ON a.OWNER = c.OWNER AND a.CONSTRAINT_NAME = c.CONSTRAINT_NAME
            WHERE c.CONSTRAINT_TYPE = 'P'
        ) pk_cols
            ON t.OWNER = pk_cols.OWNER AND t.TABLE_NAME = pk_cols.TABLE_NAME AND t.COLUMN_NAME = pk_cols.COLUMN_NAME
        LEFT JOIN (
            SELECT ic.TABLE_OWNER, ic.TABLE_NAME, ic.COLUMN_NAME, i.INDEX_NAME
            FROM ALL_IND_COLUMNS ic
            JOIN ALL_INDEXES i ON ic.INDEX_NAME = i.INDEX_NAME AND ic.TABLE_NAME = i.TABLE_NAME AND ic.TABLE_OWNER = i.OWNER
        ) idx
            ON t.OWNER = idx.TABLE_OWNER AND t.TABLE_NAME = idx.TABLE_NAME AND t.COLUMN_NAME = idx.COLUMN_NAME
        LEFT JOIN ALL_TABLES tab
    ON t.TABLE_NAME = tab.TABLE_NAME
    AND t.OWNER = tab.OWNER
    """
    params = {}
    if owner:
        query += " WHERE t.OWNER = :owner"
        params["owner"] = owner

    query += " ORDER BY t.OWNER, t.TABLE_NAME, t.COLUMN_ID"

    with connection.cursor() as cursor:
        cursor.execute(query, params)
        columns = [col[0] for col in cursor.description]
        rows = [dict(zip(columns, row)) for row in cursor.fetchall()]

    if export == "excel":
        return export_to_excel(rows, columns, filename="table_metadata.xlsx")
    
    schemas = get_schema_list()
    return render(request, "metadata/tables.html", {"data": rows, "schemas": schemas, "selected": owner})


def view_metadata(request):
    owner = request.GET.get("owner")
    export = request.GET.get("export")
    params = {"owner": owner}

    view_query = """
        SELECT v.OWNER,
               v.VIEW_NAME,
               REGEXP_SUBSTR(v.TEXT_VC, 'WHERE.+', 1,1, 'i') WHERE_CLAUSE,
               REGEXP_SUBSTR(v.TEXT_VC, 'JOIN.+?ON.+?(WHERE|GROUP BY |ORDER BBY|$)', 1, 1, 'in') as JOIN_CLAUSE

        FROM ALL_VIEWS v
        WHERE (:owner IS NULL OR v.OWNER = :owner)
    """

    dependency_query = """
        SELECT d.OWNER,
               d.NAME AS VIEW_NAME,
               LISTAGG(d.REFERENCED_OWNER || '.' || d.REFERENCED_NAME, ', ') 
                 WITHIN GROUP (ORDER BY d.REFERENCED_NAME) AS REFERENCED_OBJECTS
        FROM ALL_DEPENDENCIES d
        WHERE d.TYPE = 'VIEW' AND d.REFERENCED_TYPE IN ('TABLE', 'VIEW')
          AND (:owner IS NULL OR d.OWNER = :owner)
        GROUP BY d.OWNER, d.NAME
    """

    with connection.cursor() as cursor:
        # fetch view definitions
        cursor.execute(view_query, params)
        view_columns = [col[0] for col in cursor.description]
        views = [dict(zip(view_columns, row)) for row in cursor.fetchall()]

        # fetch referenced tables
        cursor.execute(dependency_query, params)
        dep_columns = [col[0] for col in cursor.description]
        dependencies = [dict(zip(dep_columns, row)) for row in cursor.fetchall()]

    # merge results in Python
    dep_map = {(d["OWNER"], d["VIEW_NAME"]): d["REFERENCED_OBJECTS"] for d in dependencies}
    for view in views:
        key = (view["OWNER"], view["VIEW_NAME"])
        view["REFERENCED_OBJECTS"] = dep_map.get(key, "")

    if export == "excel":
        columns = list(views[0].keys()) if views else []
        return export_to_excel(views, columns, filename="views_metadata.xlsx")

    schemas = get_schema_list()
    return render(request, "metadata/views.html", {"data": views, "schemas": schemas, "selected": owner})


def procedure_metadata(request):
    owner = request.GET.get("owner")
    export = request.GET.get("export")

    query = """
        SELECT OBJECT_NAME,
               PROCEDURE_NAME,
               OBJECT_TYPE,
               OWNER
        FROM ALL_PROCEDURES
    """
    params = {}
    if owner:
        query += " WHERE OWNER = :owner"
        params["owner"] = owner

    query += " ORDER BY OWNER, OBJECT_NAME, PROCEDURE_NAME"

    with connection.cursor() as cursor:
        cursor.execute(query, params)
        columns = [col[0] for col in cursor.description]
        rows = [dict(zip(columns, row)) for row in cursor.fetchall()]

    if export == "excel":
        return export_to_excel(rows, columns, filename="procedure_metadata.xlsx")

    schemas = get_schema_list()
    return render(request, "metadata/program_units.html", {"data": rows, "schemas": schemas, "selected": owner})


def job_metadata(request):
    owner = request.GET.get("owner")
    export = request.GET.get("export")

    query = """
        SELECT 
            OWNER,
            JOB_NAME,
            ENABLED,
            STATE,
            JOB_TYPE,
            JOB_ACTION,
            START_DATE,
            REPEAT_INTERVAL,
            NEXT_RUN_DATE,
            LAST_START_DATE,
            LAST_RUN_DURATION,
            COMMENTS
        FROM ALL_SCHEDULER_JOBS
    """
    params = {}
    if owner:
        query += " WHERE OWNER = :owner"
        params["owner"] = owner

    query += " ORDER BY OWNER, JOB_NAME"

    with connection.cursor() as cursor:
        cursor.execute(query, params)
        columns = [col[0] for col in cursor.description]
        rows = [dict(zip(columns, row)) for row in cursor.fetchall()]

    if export == "excel":
        return export_to_excel(rows, columns, filename="job_metadata.xlsx")

    schemas = get_schema_list()
    return render(request, "metadata/jobs.html", {
        "data": rows,
        "schemas": schemas,
        "selected": owner
    })


def job_run_details(request):
    owner = request.GET.get("owner")
    export = request.GET.get("export")

    query = """
        SELECT 
            OWNER,
            JOB_NAME,
            STATUS,
            ACTUAL_START_DATE,
            RUN_DURATION,
            ERROR#,
            ADDITIONAL_INFO
        FROM ALL_SCHEDULER_JOB_RUN_DETAILS
    """
    params = {}
    if owner:
        query += " WHERE OWNER = :owner"
        params["owner"] = owner

    query += " ORDER BY ACTUAL_START_DATE DESC"

    with connection.cursor() as cursor:
        cursor.execute(query, params)
        columns = [col[0] for col in cursor.description]
        rows = [dict(zip(columns, row)) for row in cursor.fetchall()]

    if export == "excel":
        return export_to_excel(rows, columns, filename="job_run_details.xlsx")

    schemas = get_schema_list()
    return render(request, "metadata/job_run_details.html", {
        "data": rows,
        "schemas": schemas,
        "selected": owner
    })


def job_logs(request):

    owner = request.GET.get("owner")
    export = request.GET.get("export")
    query = """
        SELECT 
            OWNER,
            JOB_NAME,
            LOG_DATE,
            STATUS,
            OPERATION
        FROM ALL_SCHEDULER_JOB_LOG
    """
    params = {}
    if owner:
        query += " WHERE OWNER = :owner"
        params["owner"] = owner

    query += " ORDER BY LOG_DATE DESC"

    with connection.cursor() as cursor:
        cursor.execute(query, params)
        columns = [col[0] for col in cursor.description]
        rows = [dict(zip(columns, row)) for row in cursor.fetchall()]

    if export == "excel":
        return export_to_excel(rows, columns, filename="job_logs.xlsx")

    schemas = get_schema_list()
    return render(request, "metadata/job_logs.html", {
        "data": rows,
        "schemas": schemas,
        "selected": owner
    })


def export_to_excel(data, columns, filename="output.xlsx"):
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Data"

    # نوشتن عنوان ستون‌ها
    for col_num, column_title in enumerate(columns, 1):
        col_letter = get_column_letter(col_num)
        ws[f"{col_letter}1"] = column_title

    # نوشتن داده‌ها
    for row_num, row_data in enumerate(data, 2):
        for col_num, column_title in enumerate(columns, 1):
            col_letter = get_column_letter(col_num)
            ws[f"{col_letter}{row_num}"] = row_data.get(column_title)

    response = HttpResponse(
        content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )
    response["Content-Disposition"] = f'attachment; filename="{filename}"'
    wb.save(response)
    return response