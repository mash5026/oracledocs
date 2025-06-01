from django.contrib import admin
from django.apps import apps
from import_export.admin import ImportExportModelAdmin
from django_extensions.admin import ForeignKeyAutocompleteAdmin
from .models import User, ComComplainanttype, Contracttype, WrsSubsystemsSteps, Location,Office,Complaint, ComMeeting, Profile, Profilerole, Officestaff, Gender, Committee, Committeecalendar, Committeebranch, Committeetype, Committeesupportlocation, Zone
from django.core.exceptions import ValidationError
from django.contrib import messages
from jalali_date.admin import ModelAdminJalaliMixin, TabularInlineJalaliMixin
from .forms import ProfileForm, ProfileRoleForm, OfficeForm, CommitteeForm, CommitteesupportlocationForm, ZoneForm, CommitteecalendarForm
from django.contrib.admin import SimpleListFilter
from django.http import HttpResponse
from openpyxl import Workbook
from openpyxl.utils import get_column_letter
from reportlab.pdfgen import canvas
from reportlab.lib.pagesizes import letter
from import_export.admin import ImportMixin
from django.core.cache import cache
from django.db import connection
from threading import local
from jalali_date import datetime2jalali
# Register your models here.

TABLES = [
    'Committeebranch', 'Committeetype'
]

class BaseModelAdmin(ImportExportModelAdmin, ForeignKeyAutocompleteAdmin, admin.ModelAdmin):
    list_display = []
    search_fields = []

    def __init__(self, model, admin_site):
        list_display = [field.name for field in model._meta.fields if not field.is_relation]
        search_fields = [field.name for field in model._meta.fields if not field.is_relation]
        
        if list_display:
            self.list_display = list_display
        if search_fields:
            self.search_fields = search_fields
        
        super().__init__(model, admin_site)


models = apps.get_containing_app_config("backoffice").get_models()

for model in models:
    if model.__name__ != 'Profile' and model.__name__ != 'Profilerole' and model.__name__ != 'Office' and model.__name__ != 'Committeebranch' and model.__name__ != 'Committeetype'and model.__name__ in TABLES:
        try:
            admin.site.register(model, type(f'{model.__name__}Admin', (BaseModelAdmin,), {}))
        except admin.sites.AlreadyRegistered:
            pass
        except Exception as e:
            import logging
            logger = logging.getLogger(__name__)
            logger.error(f"Error registering {model.__name__}: {e}")


class ProfileRoleInline(admin.TabularInline):
    model = Profilerole
    form = ProfileRoleForm
    fields = ['role']
    extra = 1
    can_delete = True


class ProfileOfficeInline(admin.TabularInline):
    model = Officestaff
    form = OfficeForm
    fields = ['office', 'role']
    extra = 1
    can_delete = True

    def get_formset(self, request, obj=None, **kwargs):
        formset = super().get_formset(request, obj, **kwargs)
        formset.form.current_user = request.user
        return formset


class CommitteebranchInline(admin.TabularInline):
    model = Committeebranch
    fields = ['code', 'name', 'isactive', 'version']
    extra = 1


class CommitteesupportlocationInline(admin.TabularInline):
    model = Committeesupportlocation
    form = CommitteesupportlocationForm
    fields = ['location', 'zone']
    extra = 1

    def get_formset(self, request, obj=None, **kwargs):
        formset = super().get_formset(request, obj, **kwargs)
        formset.form.current_user = request.user
        return formset


class CommitteecalendarInline(TabularInlineJalaliMixin, admin.TabularInline):
    model = Committeecalendar
    form = CommitteecalendarForm
    fields = ['branch','availabledate', 'availabletrialcount', 'maxdurationofeachmeeting', 'trialcount', 'version', 'isactive']
    extra = 1

    def get_fields(self, request, obj=None):
        if obj is None:  # If obj is None, we're adding a new instance
            return [
                'availabledate',
                'availabletrialcount',
                'maxdurationofeachmeeting',
                'trialcount',
                'version',
                'isactive'
            ]
        else:  # If obj exists, we're editing an existing instance
            return [
                'branch',
                'availabledate',
                'availabletrialcount',
                'maxdurationofeachmeeting',
                'trialcount',
                'version',
                'isactive'
            ]

    def save_model(self, request, obj, form, change):
        obj.modifiedby = request.user.id
        
        #if not obj.pk:
        if not obj.createdby:
            obj.createdby = request.user.id

        super().save_model(request, obj, form, change)

    def save_new(self, form, commit=True):
            obj = super().save_new(form, commit=False)

            # Assign the captured branch ID to the 'branch' field if it's empty
            if not obj.branch:
                current_branch_id = get_current_branch_id()
                if current_branch_id:
                    obj.branch = current_branch_id  # Directly assign the branch ID
                else:
                    raise ValueError("No branch was created to assign.")

            if commit:
                obj.save()
            return obj
    
def export_profiles_as_excel(modeladmin, request, queryset):
    # Create an HttpResponse object with content_type set to Excel file format
    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = 'attachment; filename=profiles.xlsx'

    # Create a workbook and active sheet
    wb = Workbook()
    ws = wb.active
    ws.title = "Profiles"

    # Set headers for the Excel file
    headers = ['Username', 'Full Name', 'National ID', 'Mobile Number', 'Gender', 'Location']
    ws.append(headers)

    # Iterate over the queryset and add the profile data to the sheet
    for profile in queryset:
        full_name = f"{profile.firstname} {profile.lastname}"  # Combine first and last name


        # Prepare the row of data
        row = [
            profile.username,
            full_name,
            profile.NATIONALID,
            profile.mobilenumber,
            profile.gender,
            profile.location
        ]
        ws.append(row)

    # Auto adjust column widths based on the content
    for col in range(1, len(headers) + 1):
        max_length = 0
        column = get_column_letter(col)
        for row in ws.iter_rows(min_col=col, max_col=col):
            for cell in row:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(cell.value)
                except:
                    pass
        adjusted_width = (max_length + 2)
        ws.column_dimensions[column].width = adjusted_width

    # Save the workbook to the HttpResponse object
    wb.save(response)
    return response
export_profiles_as_excel.short_description = "خروجی با فرمت اکسل"

def export_profiles_as_pdf(modeladmin, request, queryset):
    # Prepare the HTTP response with PDF content type
    response = HttpResponse(content_type='application/pdf')
    response['Content-Disposition'] = 'attachment; filename=profiles.pdf'

    # Create the PDF using ReportLab
    p = canvas.Canvas(response, pagesize=letter)
    width, height = letter

    # Set the title of the PDF
    p.setFont("Helvetica", 16)
    p.drawString(200, height - 40, "Profile Data")

    # Set the font for the content
    p.setFont("Helvetica", 12)

    # Create table headers
    p.drawString(50, height - 80, "Username")
    p.drawString(150, height - 80, "Full Name")
    p.drawString(300, height - 80, "National ID")
    p.drawString(400, height - 80, "Mobile Number")
    p.drawString(500, height - 80, "Gender")
    p.drawString(550, height - 80, "Location")

    # Iterate over the queryset and add data to the PDF
    y_position = height - 100
    for profile in queryset:
        full_name = f"{profile.firstname} {profile.lastname}"
        p.drawString(50, y_position, profile.username)
        p.drawString(150, y_position, full_name)
        p.drawString(300, y_position, str(profile.NATIONALID))
        p.drawString(400, y_position, str(profile.mobilenumber))
        p.drawString(500, y_position, str(profile.gender))
        p.drawString(550, y_position, str(profile.location))

        y_position -= 20
        if y_position < 50:
            p.showPage()  # Create a new page if content overflows
            y_position = height - 40

    p.showPage()
    p.save()
    return response
export_profiles_as_pdf.short_description = "خروجی با فرمت PDF"


class RoleFilter(SimpleListFilter):
    title = 'نقش'  # Display name in the filter sidebar
    parameter_name = 'roles'  # Query parameter in the URL

    def lookups(self, request, model_admin):
        # Provide a list of roles to filter by
        roles = Profilerole.objects.values_list('role__name', flat=True).distinct()
        return [(role, role) for role in roles]

    def queryset(self, request, queryset):
        # Apply the filter to the queryset
        if self.value():
            return queryset.filter(profileroles__role__name=self.value())
        return queryset
    

class GenderFilter(SimpleListFilter):
    title = 'جنسیت'  # Display name in the filter sidebar
    parameter_name = 'genders'  # Query parameter in the URL

    def lookups(self, request, model_admin):
        # Provide a list of roles to filter by
        names = Gender.objects.values_list('id','name')
        return [(str(g_id), g_name) for g_id, g_name in names]

    def queryset(self, request, queryset):
        # Apply the filter to the queryset
        if self.value():
            return queryset.filter(gender=self.value())
        return queryset
    
_thread_locals = local()
    
def get_current_branch_id():
    return getattr(_thread_locals, 'current_branch_id', None)

    
class ProfileAdmin(ImportMixin, admin.ModelAdmin):
    form = ProfileForm
    actions = [export_profiles_as_pdf, export_profiles_as_excel]
    #resource_class = ProfileResource

    list_display = [
        'firstname', 'lastname', 'NATIONALID', 'get_gender', 'get_loc', 'get_roles', 
        'mobilenumber', 'isactive', 'createdby', 
        'force_password_change', 'force_profile_completion'
    ]

    search_fields = [
        'firstname__icontains', 
        'lastname__icontains', 
        'NATIONALID__exact', 
        'mobilenumber__exact', 
        'username__exact'
    ]

    readonly_fields = ['createdat', 'createdby', 'modifiedat', 'modifiedby']
    
    list_filter = (RoleFilter, GenderFilter)
    list_per_page = 20
    #show_full_result_count = False
    
    inlines = [ProfileRoleInline, ProfileOfficeInline]

    class Media:
        js = ('admin_profile.js', 'nationalid_check.js',)

    def get_roles(self, obj):
        # Get all related roles, return '-' if there are none
        roles = [str(role.role) for role in obj.profileroles.all()]  # Assuming role has a `__str__` method
        return ", ".join(roles) if roles else '-'

    get_roles.short_description = 'نقش'

    def get_loc(self, obj):
        cache_key = f"location_{obj.location}"
        location_name = cache.get(cache_key)
        if not location_name:
            with connection.cursor() as cursor:
                cursor.execute("SELECT name FROM location WHERE id = %s", [obj.location])
                result = cursor.fetchone()
                location_name = result[0] if result else ''
                cache.set(cache_key, location_name, timeout=3600)  # Cache for 1 hour
        return location_name
    get_loc.short_description = 'نام شهر'

    def get_gender(self, obj):
        try:
            gen_name = Gender.objects.get(id=obj.gender).name            
        except Gender.DoesNotExist:
            gen_name = ''
        return gen_name
    get_gender.short_description = 'جنسیت'
    
    def get_form(self, request, obj=None, **kwargs):
        # Pass the current user to the form
        kwargs['form'] = ProfileForm
        form = super().get_form(request, obj, **kwargs)
        form.current_user = request.user  # Set the user
        return form
    
    # def get_queryset(self, request):

    #     qs = super().get_queryset(request)
        
    #     if request.user.is_superuser:
    #          if not hasattr(request.user, 'province') or request.user.province is None:  # Superuser can see all profiles
    #             return qs
    #          return qs.filter(location=request.user.province)
    #     elif request.user.groups.filter(name__startswith='CanCreateUser-').exists():
    #         allowed_locations = [group.name.split('-')[1] for group in request.user.groups.all() if group.name.startswith('CanCreateUser-')]
    #         return qs.filter(location__name__in=allowed_locations)
    #     return qs.none()

    def get_fields(self, request, obj=None):
        if obj:  # If obj is not None, this is an update view
            return [
                'NATIONALID', 'usertype', 'firstname', 'lastname', 'location', 
                'gender', 'mobilenumber', 'isactive', 'createdby', 
                'force_password_change', 'force_profile_completion', 'password',
                'deathdate', 'activefrom', 'activeto', 'childrennumber',
                'createdat', 'defunctdate', 'foundationdate', 'graduation', 
                'maritalstatus', 'modifiedat', 'modifiedby', 'religion', 
                'postalcode', 'zonecode', 'faxnumber', 
                'nationalcertificatecode', 'nationalcertificateserial', 
                'phonenumber', 'fathername', 'organizationname', 'username', 
                'address', 'secretkey', 'birthdate', 'currenttaminworkshopcode', 
                'email', 'graduationlevel', 'physicalcondition', 'idno', 
                'ispermittedtoreceivebyemail', 'ispermittedtoreceivebyfax', 
                'militaryservicestatus', 'parentlocation', 'nationality', 
                'officialemail', 'externalusertype', 'is_converted', 
                'birth_place'
            ]
        else:

            return [
                'NATIONALID', 'usertype', 'firstname', 'lastname', 'location', 
                'gender', 'mobilenumber', 'isactive', 'createdby', 
                'force_password_change', 'force_profile_completion', 'password'
            ]        
       
    def save_model(self, request, obj, form, change):
        """
        Override the save_model method to catch validation errors from the model
        and prevent success messages when there are errors.
        """
        try:
            if not obj.pk:  # This means it's a new object being created
                if not obj.createdby:
                    obj.createdby = request.user.id  # Set createdby to the current user

            obj.modifiedby = request.user.id

            # Attempt to save the object
            super().save_model(request, obj, form, change)

        except ValidationError as e:
            # Catch validation errors and show them on the form
            form.add_error(None, e)  # Adds the error to the form's non-field errors
            messages.set_level(request, messages.ERROR)
            messages.add_message(request, messages.ERROR, messages.error(request, e.message))  # Optionally show a Django message
            return

    def save_formset(self, request, form, formset, change):
        """
        Override save_formset to handle deletions explicitly.
        """
        if formset.model == Profilerole or formset.model == Officestaff:
            # Check for any deletions
            instances = formset.save(commit=False)
            for obj in formset.deleted_objects:
                obj.delete()  # Handle deletion
            for instance in instances:
                instance.save()
        else:
            super().save_formset(request, form, formset, change)

class CommitteeAdmin(ImportExportModelAdmin, ModelAdminJalaliMixin, admin.ModelAdmin):
    form = CommitteeForm
    list_display = ['code', 'name', 'commiteetype', 'office', 'address', 'createdat', 'createdby', 'isactive',
                    'modifiedat', 'modifiedby',  'version']
    readonly_fields = ['createdat', 'createdby', 'modifiedat', 'modifiedby']
    list_display_links = ["code","name","commiteetype", "office"]
    list_filter = ('name', 'commiteetype')
    list_per_page = 30

    inlines = [CommitteebranchInline, CommitteecalendarInline, CommitteesupportlocationInline]

    def get_form(self, request, obj=None, **kwargs):
        
        kwargs['form'] = CommitteeForm
        form = super().get_form(request, obj, **kwargs)
        form.current_user = request.user  # Set the user
        return form
   
    def save_model(self, request, obj, form, change):
        if form.cleaned_data.get('office'):
            obj.office = form.cleaned_data['office']
        obj.modifiedby = request.user.id
        if not obj.pk:
            if not obj.createdby:
                    obj.createdby = request.user.id
            
        super().save_model(request, obj, form, change)


class CommitteebranchAdmin(admin.ModelAdmin):
    list_display = ['code', 'name', 'committee', 'get_comms' ,'createdat', 'createdby', 'isactive', 'modifiedat', 'modifiedby', 'version']

    def get_comms(self, obj):
        # Get all related roles, return '-' if there are none
        comms = [str(comm.name) for comm in obj.Committee.all()]  # Assuming role has a `__str__` method
        return ", ".join(comms) if comms else '-'

    get_comms.short_description = 'کمیته'

    def save_model(self, request, obj, form, change):
        obj.modifiedby = request.user.id

        if not obj.pk:
            if not obj.createdby:
                    obj.createdby = request.user.id
        super().save_model(request, obj, form, change)


class CommitteecalendarAdmin(ModelAdminJalaliMixin, admin.ModelAdmin):

    form = CommitteecalendarForm
    list_display = ['committee', 'branch', 'get_availabledate_jalali', 'availabletrialcount', 'maxdurationofeachmeeting', 'createdat', 'createdby', 
                    'modifiedat', 'modifiedby', 'trialcount', 'version', 'isactive']

    
    @admin.display(description='تاریخ های دردسترس', ordering='createdat')
    def get_availabledate_jalali(self, obj):
            return datetime2jalali(obj.availabledate).strftime('%a, %d %b %Y %H:%M:%S')
    

    def save_model(self, request, obj, form, change):
        obj.modifiedby = request.user.id        
        #if not obj.pk:
        if not obj.createdby:
            obj.createdby = request.user.id
        super().save_model(request, obj, form, change)


class CommitteesupportlocationAdmin(admin.ModelAdmin):

    form = CommitteesupportlocationForm

    def get_form(self, request, obj=None, **kwargs):
        
        kwargs['form'] = CommitteesupportlocationForm
        form = super().get_form(request, obj, **kwargs)
        form.current_user = request.user  # Set the user
        return form

    def save_model(self, request, obj, form, change):
        obj.modifiedby = request.user.id
        
        if not obj.pk:
            if not obj.createdby:
                    obj.createdby = request.user.id
        super().save_model(request, obj, form, change)


class ZoneAdmin(admin.ModelAdmin):

    list_display = [
        'code', 'name', 'location'
        ]

    form=ZoneForm

    def get_form(self, request, obj=None, **kwargs):
        
        kwargs['form'] = ZoneForm
        form = super().get_form(request, obj, **kwargs)
        form.current_user = request.user  # Set the user
        return form

admin.site.register(Profile, ProfileAdmin)
admin.site.register(Committee, CommitteeAdmin)
admin.site.register(Committeecalendar, CommitteecalendarAdmin)
admin.site.register(Committeesupportlocation, CommitteesupportlocationAdmin)
admin.site.register(Zone, ZoneAdmin)
